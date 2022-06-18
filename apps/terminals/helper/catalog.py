from datetime import datetime
from io import BytesIO
from typing import ClassVar, Dict

from django.db import transaction
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.exceptions import ValidationError
from rest_framework.serializers import Serializer

from apps.terminals.models import CatalogImport
from apps.terminals.models.catalog import ImportStatus
from apps.terminals.serializers.product import AddProductSerializer
from config.settings.dev import DATETIME_FORMAT, USE_LOCAL_WORKBOOK
from core import external_apis
from core.exception import ImportException
from core.utils import is_dict_values_none, iterate_all

ERROR_MSG = 'Lỗi dữ liệu, tải file lỗi để xem chi tiết'
ERROR_FONT = Font(color="FF0000", name='Consolas')
SHEET_MIME_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def get_by_catalog_import_input(queryset, request_data):
    ids = request_data.get('ids')
    if ids:
        queryset = queryset.filter(id__in=ids)

    started_at_from = request_data.get('started_at_from')
    if started_at_from is not None:
        queryset.filter(created_at__gte=started_at_from)

    ended_at_to = request_data.get('ended_at_to')
    if ended_at_to is not None:
        queryset.filter(created_at__lte=ended_at_to)

    user_id = request_data.get('user_id')
    if user_id is not None:
        queryset.filter(user_id=user_id)

    status = request_data.get('status')
    if status is not None:
        queryset.filter(status=status)

    return queryset


class ImportProductHandler:

    def __init__(self, catalog: CatalogImport):
        self.catalog = catalog
        self.user = catalog.user
        self.row_index = 4
        self.max_quantity = 100
        self.total_rows = 0
        self.serializer_cls = AddProductSerializer
        self.errors: Dict[int, str] = {}
        self.error_column = 'F'
        self.file_url = catalog.source_file
        self.max_try = 5
        self.trying = 0

    @staticmethod
    def convert_cell_type(cell: Cell):
        value = cell.value
        if value and isinstance(value, str):
            value = value.strip()
        if isinstance(value, datetime):
            value = value.strftime(DATETIME_FORMAT)
        return value

    def get_work_book(self, file_url=None, data_only=True):
        if USE_LOCAL_WORKBOOK:
            return load_workbook(file_url or self.file_url or self.catalog.source_file, data_only=data_only)
        file = external_apis.download_file(file_url or self.file_url or self.catalog.source_file)
        return load_workbook(file, data_only=data_only)

    def get_data_row_index(self):
        return self.row_index

    def get_max_quantity(self):
        return self.max_quantity

    def get_serializer_cls(self):
        return self.serializer_cls

    def get_error_column(self):
        return self.error_column

    def extract_data(self, **kwargs):
        self.catalog.status = ImportStatus.EXTRACT
        self.catalog.save()
        workbook = self.get_work_book()
        sheet = workbook.worksheets[0]
        index = 0
        data = []
        for row in sheet.rows:
            index += 1
            if index < self.get_data_row_index():
                continue
            serializer_cls = self.get_serializer_cls()
            keys = serializer_cls().get_fields().keys()
            row_data = {key: self.convert_cell_type(row[index]) for index, key in enumerate(keys)}
            if not is_dict_values_none(row_data):
                row_data['row_index'] = index
                data.append(row_data)
                self.trying = 0
            else:
                self.trying += 1
                if self.trying > self.max_try:
                    break
        if len(data) > self.get_max_quantity():
            raise ValueError(f'Số lượng import > {self.get_max_quantity()}')
        if len(data) == 0:
            raise ValueError('File không có dữ liệu import')
        self.total_rows = len(data)
        return data

    def process(self):
        import_data = self.extract_data()
        self.catalog.status = ImportStatus.PROCESS
        self.catalog.save()
        exception = None
        try:
            with transaction.atomic():
                sid = transaction.savepoint()
                serializer_cls: ClassVar[Serializer] = self.get_serializer_cls()
                for row in import_data:
                    try:
                        row_index = row.get('row_index')
                        row_serializer = serializer_cls(data=row, context={'user': self.user})
                        row_serializer.is_valid(raise_exception=True)
                        row_serializer.save()
                    except Exception as err:
                        err_str = str(err)
                        if isinstance(err, ValidationError):
                            err_str = '\n'.join(
                                [str(item) for item in list(iterate_all(err.get_full_details(),  # type: ignore
                                                                        ignore_keys=['code'],
                                                                        returned="value"))])
                        self.errors[row_index] = err_str
                if len(self.errors) > 0:
                    transaction.savepoint_rollback(sid)
                    all_errors = {**self.errors}
                    message = ERROR_MSG
                    raise ImportException(message, errors=all_errors)

        except Exception as err:
            self.catalog.status = ImportStatus.FAIL
            if isinstance(err, ImportException):
                try:
                    error_file_url = self.handle_error(err.errors)
                    self.catalog.result_file = error_file_url
                except Exception as ex:
                    pass
            else:
                exception = err
            self.catalog.save()
        finally:
            if not (self.errors or exception):
                self.catalog.status = ImportStatus.SUCCESS
                self.catalog.quantity_product = self.total_rows
                self.catalog.save()
            else:
                raise ValidationError(message or str(exception))

    def handle_error(self, errors):
        workbook = self.get_work_book(file_url=self.catalog.source_file)
        sheet = workbook.worksheets[0]
        # self.clear_old_error_data(sheet)
        # if not self.workflow.error_file_url:
        #     self.style_error_header(sheet)
        error_column = self.get_error_column()
        for row_index, err_string in errors.items():
            cell_id = f'{error_column}{row_index}'
            sheet[cell_id] = err_string
            sheet[cell_id].font = ERROR_FONT

        files = {'file': (f'error_{self.catalog.user.name}',
                          BytesIO(save_virtual_workbook(workbook)),
                          SHEET_MIME_TYPE)}
        error_file_url = external_apis.upload_file(files)
        return error_file_url
